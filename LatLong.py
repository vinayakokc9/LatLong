import requests
from bs4 import BeautifulSoup

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import re

def _getLatLong(cityState):
    url = "https://www.travelmath.com/cities/" 
    invalidLocation = "find a city map" # used for invalid input
       
    if(cityState == "None"):
        return "", ""
    
    # example format "New York, NY"
    cityState = cityState.split(',')
    if(len(cityState) != 2):
        return "", ""
    
    cityState[0] = cityState[0].strip()
    cityState[0] = cityState[0].replace(" ", "+")
    
    cityState[1] = cityState[1].strip()
    cityState[1] = "+" + cityState[1]

    #creating complete url
    url = url + cityState[0] + "," + cityState[1] 
    print(url)

    req = requests.get(url)
    soup = BeautifulSoup(req.content, "html.parser")
    
    # example format 42° 1' 51" N / 93° 37' 54" W
    latLongString = soup.h3.get_text() 
    latLongString = latLongString.split("/")
    
    # handles invalid input and location
    if(latLongString.lower() == invalidLocation.lower()):
        return "", ""
    
    # removing compass directions 
    latVal = latLongString[0][:-3]

    # removing space at the beginning and compass direction
    longVal = latLongString[1][1:-2]

    return latVal, longVal
    


def autopopulate(filePath):

    targetCol = 2 # column 2 contains city and state values
    startingRow = 2 # ignoring header
    latColumn = 3
    longColumn = 4
    targetDatasheet = 'Datasheet'

    # setting up Excel file for read and write operations
    wb = load_workbook(filePath, read_only=False, keep_vba=True)
    ws = wb[targetDatasheet]

    for row in ws.iter_rows(min_row=startingRow, min_col=targetCol, max_col=targetCol):
        for cell in row:
            if(cell.value is None):
                break
            lat, long = _getLatLong(cell.value)
            print(f'{cell.row} {cell.column} {cell.value} {lat} {long}')
            ws.cell(row=cell.row, column=latColumn, value=lat)
            ws.cell(row=cell.row, column=longColumn, value=long)

    wb.save(filePath)
        
        
        
        
        
        
        
        
        
        
        
        